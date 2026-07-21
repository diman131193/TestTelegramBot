from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR))

from app.diagnostics import build_recommendation_with_sources  # noqa: E402
from app.diagnostics_validation import ValidationReport, validate_sources  # noqa: E402

DATA_DIR = ROOT_DIR / "data"
DEFAULT_WORKBOOK = ROOT_DIR / "content" / "diagnostics-content.xlsx"
ROLE_TO_INTERNAL = {
    "Основная": "primary",
    "Стоп-флаг": "alert",
    "Дополнение": "addon",
    "primary": "primary",
    "alert": "alert",
    "addon": "addon",
}
NUMBER_SETTINGS = {"max_addons", "max_actions", "max_cautions"}


def _load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as file:
        value = json.load(file)
    if not isinstance(value, dict):
        raise ValueError(f"{path.name}: ожидается JSON-объект")
    return value


def _open_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Для импорта таблицы установите зависимости: pip install -r requirements.txt"
        ) from error
    return load_workbook(path, data_only=False)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = _text(value).casefold()
    if normalized in {"да", "yes", "true", "1", "активно", "готово"}:
        return True
    if normalized in {"нет", "no", "false", "0", "выключено"}:
        return False
    return default


def _sheet_rows(workbook, sheet_name: str, required_header: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"В таблице отсутствует лист «{sheet_name}».")
    sheet = workbook[sheet_name]
    header_row = None
    headers: list[str] = []
    for row_number in range(1, min(sheet.max_row, 15) + 1):
        values = [_text(cell.value) for cell in sheet[row_number]]
        if required_header in values:
            header_row = row_number
            headers = values
            break
    if header_row is None:
        raise ValueError(f"На листе «{sheet_name}» не найден столбец «{required_header}».")

    rows: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=header_row + 1, max_col=len(headers)):
        values = [cell.value for cell in cells]
        if not any(value is not None and _text(value) for value in values):
            continue
        rows.append({header: value for header, value in zip(headers, values) if header})
    return rows


def compile_workbook(path: Path) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
    workbook = _open_workbook(path)
    errors: list[str] = []

    settings: dict[str, Any] = {}
    for row in _sheet_rows(workbook, "Общие тексты", "Ключ"):
        key = _text(row.get("Ключ"))
        if not key:
            continue
        value = row.get("Значение")
        if key in NUMBER_SETTINGS:
            try:
                value = int(value)
            except (TypeError, ValueError):
                errors.append(f"Настройка {key} должна быть целым числом.")
                continue
        else:
            value = _text(value)
        settings[key] = value

    facts: dict[str, dict[str, Any]] = OrderedDict()
    for row in _sheet_rows(workbook, "Подписи", "Группа ID"):
        group_id = _text(row.get("Группа ID"))
        value_id = _text(row.get("Значение ID"))
        phrase = _text(row.get("Фраза для клиента"))
        if not group_id or not value_id:
            errors.append("На листе «Подписи» есть строка без служебного ID.")
            continue
        group = facts.setdefault(
            group_id,
            {
                "title": _text(row.get("Раздел")) or group_id,
                "visible": _bool(row.get("Показывать"), True),
                "values": OrderedDict(),
            },
        )
        if value_id in group["values"]:
            errors.append(f"Подпись {group_id}/{value_id} повторяется.")
        group["values"][value_id] = phrase

    modules: list[dict[str, Any]] = []
    for row in _sheet_rows(workbook, "Рекомендации", "ID модуля"):
        module_id = _text(row.get("ID модуля"))
        if not module_id:
            errors.append("На листе «Рекомендации» есть строка без ID модуля.")
            continue
        role_source = _text(row.get("Роль"))
        role = ROLE_TO_INTERNAL.get(role_source)
        if not role:
            errors.append(f"У рекомендации {module_id} неизвестная роль «{role_source}».")
            role = role_source
        active = _bool(row.get("Активно"), True)
        status = _text(row.get("Статус"))
        if active and status.casefold() not in {"готово", "ready"}:
            errors.append(f"Активная рекомендация {module_id} не имеет статуса «Готово».")
        actions = [
            _text(row.get(f"Шаг {number}"))
            for number in range(1, 4)
            if _text(row.get(f"Шаг {number}"))
        ]
        cautions = [
            _text(row.get(f"Осторожно {number}"))
            for number in range(1, 3)
            if _text(row.get(f"Осторожно {number}"))
        ]
        modules.append(
            {
                "id": module_id,
                "role": role,
                "situation": _text(row.get("Ситуация")),
                "title": _text(row.get("Заголовок")),
                "summary": _text(row.get("Краткое объяснение")),
                "actions": actions,
                "cautions": cautions,
                "cta": _text(row.get("CTA")),
                "service_id": _text(row.get("Услуга ID")),
                "active": active,
            }
        )

    services: list[dict[str, Any]] = []
    for row in _sheet_rows(workbook, "Услуги", "ID услуги"):
        service_id = _text(row.get("ID услуги"))
        if not service_id:
            continue
        services.append(
            {
                "id": service_id,
                "name": _text(row.get("Название")),
                "fits": _text(row.get("Кому подходит")),
                "restrictions": _text(row.get("Когда нужна консультация")),
                "cta": _text(row.get("CTA")),
                "url": _text(row.get("Ссылка")),
                "active": _bool(row.get("Активно"), True),
            }
        )

    question_rows = _sheet_rows(workbook, "Вопросы", "ID вопроса")
    questions: dict[str, dict[str, Any]] = OrderedDict()
    for row in question_rows:
        if not _bool(row.get("Активно"), True):
            continue
        question_id = _text(row.get("ID вопроса"))
        answer_id = _text(row.get("ID ответа"))
        if not question_id or not answer_id:
            errors.append("На листе «Вопросы» есть строка без ID вопроса или ответа.")
            continue
        question_text = _text(row.get("Текст вопроса"))
        question = questions.setdefault(question_id, {"text": question_text, "options": []})
        if question["text"] != question_text:
            errors.append(f"Для вопроса {question_id} указаны разные тексты.")
        question["options"].append(
            {
                "id": answer_id,
                "text": _text(row.get("Текст ответа")),
                "next": _text(row.get("Следующий вопрос")),
            }
        )

    start = settings.pop("start_question_id", None)
    content = {
        "version": 1,
        "settings": settings,
        "facts": facts,
        "modules": modules,
        "services": services,
    }
    questions_config = {"start": str(start) if start is not None else "", "questions": questions}
    return content, questions_config, errors


def validate_workbook(path: Path) -> tuple[ValidationReport, dict[str, Any], dict[str, Any]]:
    content, questions, compile_errors = compile_workbook(path)
    factors = _load_json(DATA_DIR / "diagnostic_factors.json")
    rules = _load_json(DATA_DIR / "diagnostic_rules.json")
    report = validate_sources(questions, factors, rules, content)
    report.errors[:0] = compile_errors
    return report, content, questions


def _atomic_json_write(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", dir=path.parent, delete=False, suffix=".tmp"
    ) as file:
        json.dump(value, file, ensure_ascii=False, indent=2)
        file.write("\n")
        temp_path = Path(file.name)
    os.replace(temp_path, path)
    os.chmod(path, 0o644)


def _print_report(report: ValidationReport) -> None:
    for warning in report.warnings:
        print(f"ПРЕДУПРЕЖДЕНИЕ: {warning}")
    for error in report.errors:
        print(f"ОШИБКА: {error}")
    if report.ok:
        print(
            f"OK: {report.terminal_paths} путей, самый длинный результат — "
            f"{report.longest_result} символов."
        )


def command_validate(path: Path) -> int:
    report, _, _ = validate_workbook(path)
    _print_report(report)
    return 0 if report.ok else 1


def command_preview(path: Path, answers_source: str | None) -> int:
    report, content, _ = validate_workbook(path)
    _print_report(report)
    if not report.ok:
        return 1
    factors = _load_json(DATA_DIR / "diagnostic_factors.json")
    rules = _load_json(DATA_DIR / "diagnostic_rules.json")
    answers = (
        [item.strip() for item in answers_source.split(",") if item.strip()]
        if answers_source
        else ["1.4", "2.3", "3.4", "4.2", "6.2"]
    )
    print("\n--- ПРЕДПРОСМОТР ---\n")
    print(build_recommendation_with_sources(answers, factors, rules, content))
    return 0


def command_publish(path: Path) -> int:
    report, content, questions = validate_workbook(path)
    _print_report(report)
    if not report.ok:
        print("Публикация отменена: исправьте ошибки в таблице.")
        return 1
    snapshot = {
        "version": 1,
        "published_at": datetime.now(timezone.utc).isoformat(),
        "questions": questions,
        "content": content,
    }
    _atomic_json_write(DATA_DIR / "diagnostics_snapshot.json", snapshot)
    print("Контент опубликован. Перезапуск бота не требуется для новых прохождений.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Проверка и публикация контента диагностики")
    parser.add_argument(
        "command", choices=("validate", "preview", "publish"), help="Действие"
    )
    parser.add_argument(
        "workbook", nargs="?", type=Path, default=DEFAULT_WORKBOOK, help="Путь к XLSX"
    )
    parser.add_argument(
        "--answers",
        help="ID ответов для preview через запятую, например 1.1,2.1,3.1,4.1,5.1,8.4",
    )
    args = parser.parse_args()
    path = args.workbook.resolve()
    if not path.exists():
        print(f"Файл не найден: {path}")
        return 1
    try:
        if args.command == "validate":
            return command_validate(path)
        if args.command == "preview":
            return command_preview(path, args.answers)
        return command_publish(path)
    except (OSError, ValueError, RuntimeError) as error:
        print(f"ОШИБКА: {error}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
